import psycopg2, os, io, csv, re, json, zipfile, uuid
from flask import Flask, render_template_string, request, redirect, url_for, Response, flash, session, Blueprint, send_from_directory
from openpyxl import Workbook, load_workbook
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from datetime import datetime
from functools import wraps

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', os.urandom(24))
UPLOAD_FOLDER = os.environ.get('UPLOAD_FOLDER', '/app/uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

url_prefix = os.environ.get('URL_PREFIX', '/datahub')
blueprint = Blueprint('main', __name__, url_prefix=url_prefix)

# ==================== БАЗОВЫЕ ФУНКЦИИ ====================
def get_connection():
    conn = psycopg2.connect(host=os.environ['DB_HOST'], database=os.environ['DB_NAME'],
                            user=os.environ['DB_USER'], password=os.environ['DB_PASSWORD'])
    conn.autocommit = False
    return conn

def sanitize_name(name):
    name = re.sub(r'[^a-zA-Z0-9_]', '_', name)
    if name and name[0].isdigit(): name = 't_' + name
    return name.lower()[:63]

def get_numeric_columns(cur, tablename):
    try:
        cur.execute('SELECT * FROM "' + tablename + '" LIMIT 1;')
        cols = []
        for desc in cur.description:
            if desc[0] == 'id': continue
            try:
                cur.execute('SELECT "' + desc[0] + '" FROM "' + tablename + '" WHERE "' + desc[0] + '" IS NOT NULL LIMIT 5;')
                for (v,) in cur.fetchall(): float(str(v).replace(',', '.'))
                cols.append(desc[0])
            except: pass
        return cols
    except: return []

def _compute_stats(cur, tablename, columns, total_rows):
    stats = [{'title': 'Total Rows', 'value': total_rows}, {'title': 'Columns', 'value': len(columns)}]
    for col in columns[:3]:
        try:
            cur.execute('SELECT COUNT(DISTINCT "' + col + '") FROM "' + tablename + '";')
            stats.append({'title': 'Unique: ' + col[:15], 'value': cur.fetchone()[0]})
        except: pass
    return stats[:5]

def log_action(username, action, table_name='', details=''):
    try:
        conn = get_connection(); cur = conn.cursor()
        cur.execute("INSERT INTO _logs (username, action, table_name, details) VALUES (%s,%s,%s,%s)",
                    (username, action, table_name, details[:500] if details else ''))
        conn.commit(); cur.close(); conn.close()
    except: pass

def api_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        api_key = request.headers.get('X-API-Key') or request.args.get('api_key')
        if not api_key: return json.dumps({"error": "API key required"}), 401, {'Content-Type': 'application/json'}
        conn = get_connection(); cur = conn.cursor()
        cur.execute("SELECT id FROM _users WHERE password=%s", (api_key,))
        user = cur.fetchone(); cur.close(); conn.close()
        if not user: return json.dumps({"error": "Invalid API key"}), 403, {'Content-Type': 'application/json'}
        return f(*args, **kwargs)
    return decorated

# ==================== СТИЛИ ====================
BASE_STYLE = """<style>:root{--bg:#0d1117;--surface:#161b22;--surface2:#21262d;--border:#30363d;--text:#e6edf3;--text2:#8b949e;--blue:#58a6ff;--green:#3fb950;--red:#f85149;--orange:#d2991d;--purple:#bc8cff;}.light{--bg:#fff;--surface:#f6f8fa;--surface2:#eaeef2;--border:#d0d7de;--text:#1f2328;--text2:#656d76;}*{box-sizing:border-box;margin:0;padding:0}body{font-family:-apple-system,BlinkMacSystemFont,sans-serif;background:var(--bg);color:var(--text)}.header{background:var(--surface);border-bottom:1px solid var(--border);padding:16px 24px;display:flex;justify-content:space-between}.header h1{font-size:22px;color:var(--blue)}.container{max-width:1500px;margin:0 auto;padding:24px}.card{background:var(--surface);border:1px solid var(--border);border-radius:8px;padding:20px;margin-bottom:20px}h3{font-size:16px;margin-bottom:14px}.row{display:flex;gap:16px;flex-wrap:wrap}.col{flex:1;min-width:200px}.col-2{flex:2;min-width:300px}input,textarea,select,button{font-family:inherit;font-size:14px;border-radius:6px}textarea,input:not([type=file]),select{background:var(--bg);border:1px solid var(--border);color:var(--text);padding:8px 12px;width:100%}.btn{padding:8px 16px;border:none;cursor:pointer;font-weight:600;font-size:13px;text-decoration:none;display:inline-block;border-radius:6px}.btn-primary{background:var(--blue);color:#000}.btn-green{background:var(--green);color:#000}.btn-red{background:var(--red);color:#fff}.btn-orange{background:var(--orange);color:#000}.btn-purple{background:var(--purple);color:#000}.btn-sm{padding:5px 10px;font-size:11px}.tabs{display:flex;gap:4px;margin-bottom:16px;flex-wrap:wrap}.tabs a{padding:8px 16px;background:var(--surface2);color:var(--text2);border-radius:6px;text-decoration:none}.tabs a:hover{background:var(--blue);color:#000}table{width:100%;border-collapse:collapse;font-size:13px}th,td{padding:8px 12px;text-align:left;border-bottom:1px solid var(--border)}th{background:var(--surface2);cursor:pointer}.scrollable{max-height:500px;overflow:auto;border:1px solid var(--border);border-radius:6px}.badge{background:var(--blue);color:#000;padding:2px 8px;border-radius:12px;font-size:11px}.flex-between{display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:10px}.stat-box{background:var(--surface2);border:1px solid var(--border);border-radius:8px;padding:16px;text-align:center;min-width:120px;flex:1}.stat-box .num{font-size:28px;font-weight:700}.stat-box .label{font-size:12px;color:var(--text2)}.report-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(300px,1fr));gap:16px}.chart-container{position:relative;height:300px}.nav{display:flex;gap:8px;margin-bottom:16px;flex-wrap:wrap}.nav a{padding:8px 16px;background:var(--surface2);color:var(--text2);border-radius:6px;text-decoration:none}.nav a:hover{background:var(--purple);color:#fff}.flash{padding:12px 16px;margin-bottom:16px;border-radius:6px}.flash-success{background:#1a3d2b;color:#7ee787}.flash-error{background:#3d1a1a;color:#fdaeb7}.modal-overlay{display:none;position:fixed;top:0;left:0;width:100%;height:100%;background:rgba(0,0,0,0.7);z-index:1000;justify-content:center;align-items:center}.modal-overlay.active{display:flex}.modal{background:var(--surface);border:1px solid var(--border);border-radius:8px;padding:24px;width:90%;max-width:600px}.modal h3{margin-bottom:16px}.modal label{display:block;margin:8px 0 4px;font-size:13px;color:var(--text2)}.modal input,.modal select{margin-bottom:8px}.modal .btn-row{display:flex;gap:8px;margin-top:16px;justify-content:flex-end}.filter-bar{display:flex;gap:8px;align-items:center;flex-wrap:wrap;margin-bottom:12px}.filter-bar input,.filter-bar select{width:auto;min-width:150px}.theme-toggle{cursor:pointer;background:var(--surface2);border:1px solid var(--border);color:var(--text);padding:8px 12px;border-radius:6px;font-size:18px}.photo-thumb{max-width:60px;max-height:60px;border-radius:4px;cursor:pointer;margin-right:4px}@media(max-width:768px){.row{flex-direction:column}}</style>"""
THEME_SCRIPT = """<script>document.documentElement.className=localStorage.getItem('theme')||'dark';function toggleTheme(){var t=document.documentElement.className==='dark'?'light':'dark';document.documentElement.className=t;localStorage.setItem('theme',t);document.getElementById('themeBtn').textContent=t==='dark'?'☀️':'🌙'}</script>"""

def hdr(user):
    nav = '<a href="' + url_prefix + '/">Data</a> '
    nav += '<a href="' + url_prefix + '/overview">Overview</a> '
    nav += '<a href="' + url_prefix + '/dashboard">Dashboard</a> '
    nav += '<a href="' + url_prefix + '/pivot">Pivot</a> '
    nav += '<a href="' + url_prefix + '/compare">Compare</a> '
    nav += '<a href="' + url_prefix + '/templates">Templates</a> '
    nav += '<a href="' + url_prefix + '/logs">Logs</a> '
    nav += '<a href="' + url_prefix + '/backup" class="btn btn-orange btn-sm">Backup</a>'
    return '<div class="header"><h1>Data Hub Pro v3.4</h1><div style="display:flex;align-items:center;gap:12px;"><button id="themeBtn" class="theme-toggle" onclick="toggleTheme()">☀️</button><span style="color:var(--text2)">' + user + '</span><a href="' + url_prefix + '/logout" class="btn btn-red btn-sm">Logout</a></div></div><div class="container"><div class="nav">' + nav + '</div>'
# ==================== LOGIN ====================
LOGIN_HTML = f"""<!DOCTYPE html><html><head><meta charset="UTF-8"><title>Data Hub Pro v3.4</title><style>body{{font-family:sans-serif;background:#0d1117;color:#e6edf3;display:flex;justify-content:center;align-items:center;height:100vh;margin:0}}.box{{background:#161b22;border:1px solid #30363d;border-radius:8px;padding:24px;width:400px}}h3{{color:#58a6ff;margin-bottom:16px}}input{{width:100%;padding:10px;background:#0d1117;border:1px solid #30363d;color:#e6edf3;border-radius:6px;margin-bottom:12px}}.btn{{width:100%;padding:10px;background:#58a6ff;border:none;border-radius:6px;font-weight:bold;cursor:pointer;color:#000}}.err{{background:#3d1a1a;color:#fdaeb7;padding:8px;border-radius:4px;margin-bottom:12px}}</style></head><body><div class="box"><h3>Data Hub Pro v3.4</h3>FLASH<form method="POST"><input name="username" placeholder="Username"><input type="password" name="password" placeholder="Password"><button class="btn">Login</button></form><p style="font-size:12px;color:#8b949e;margin-top:12px;">Default: admin / your_password</p></div></body></html>"""

@blueprint.route('/login', methods=['GET', 'POST'])
def login_page():
    if request.method == 'POST':
        u = request.form.get('username', ''); p = request.form.get('password', '')
        conn = get_connection(); cur = conn.cursor()
        cur.execute("SELECT id, role FROM _users WHERE username=%s AND password=%s", (u, p))
        user = cur.fetchone(); cur.close(); conn.close()
        if user: session['user'] = u; session['role'] = user[1]; log_action(u, 'LOGIN'); return redirect(url_for('main.index'))
        flash('Invalid credentials!', 'error')
    from flask import get_flashed_messages
    msgs = get_flashed_messages(with_categories=True)
    flash_html = ''.join(['<div class="err">' + m + '</div>' for c, m in msgs])
    return LOGIN_HTML.replace('FLASH', flash_html)

@blueprint.route('/logout')
def logout():
    if session.get('user'): log_action(session['user'], 'LOGOUT')
    session.clear(); return redirect(url_for('main.login_page'))

# ==================== ГЛАВНАЯ ====================
@blueprint.route('/')
def index():
    if not session.get('user'): return redirect(url_for('main.login_page'))
    conn = get_connection(); cur = conn.cursor()
    cur.execute("SELECT table_name FROM _meta ORDER BY created_at DESC;"); tables = cur.fetchall()
    cur.close(); conn.close()
    tabs = ''.join(['<a href="' + url_prefix + '/table/' + t[0] + '">' + t[0] + '</a>' for t in tables])
    return BASE_STYLE + THEME_SCRIPT + hdr(session['user']) + '<div class="card"><h3>Import Data</h3><form method="POST" action="' + url_prefix + '/import" enctype="multipart/form-data"><div class="row"><div class="col-2"><input type="file" name="datafile" accept=".csv,.xlsx,.xls" required></div><div class="col"><input type="text" name="tablename" placeholder="Table name"></div><div><button class="btn btn-primary">Upload & Load</button></div></div></form></div><div class="tabs">' + tabs + '</div></div></body></html>'

# ==================== IMPORT ====================
@blueprint.route('/import', methods=['POST'])
def import_data():
    if not session.get('user'): return redirect(url_for('main.login_page'))
    file = request.files.get('datafile')
    if not file or file.filename == '': flash('No file!','error'); return redirect(url_for('main.index'))
    tablename = sanitize_name(request.form.get('tablename','').strip() or os.path.splitext(file.filename)[0])
    ext = os.path.splitext(file.filename)[1].lower()
    try:
        if ext == '.csv':
            stream = io.StringIO(file.read().decode('utf-8')); reader = csv.DictReader(stream)
            fieldnames = list(reader.fieldnames); rows_data = [dict(row) for row in reader]
        elif ext in ['.xlsx','.xls']:
            wb = load_workbook(file); ws = wb.active; rows_iter = ws.iter_rows(values_only=True)
            fieldnames = [str(c) if c else 'col_'+str(i) for i,c in enumerate(next(rows_iter))]
            rows_data = [{fieldnames[i]: str(row[i]) if row[i] is not None else '' for i in range(len(fieldnames))} for row in rows_iter if any(c is not None for c in row)]
        else: flash('Unsupported format!','error'); return redirect(url_for('main.index'))
        conn = get_connection(); cur = conn.cursor()
        cur.execute('CREATE TABLE IF NOT EXISTS "' + tablename + '" (id SERIAL PRIMARY KEY, ' + ', '.join(['"' + c + '" VARCHAR(1000)' for c in fieldnames]) + ');')
        cur.execute("INSERT INTO _meta (table_name) VALUES (%s) ON CONFLICT DO NOTHING;", (tablename,))
        conn.commit()
        for row in rows_data:
            cur.execute('INSERT INTO "' + tablename + '" (' + ','.join(['"' + c + '"' for c in fieldnames]) + ') VALUES (' + ','.join(['%s']*len(fieldnames)) + ');', [row.get(c,'') for c in fieldnames])
        conn.commit(); cur.close(); conn.close(); log_action(session['user'], 'IMPORT', tablename, str(len(rows_data))+' rows')
        flash('Loaded ' + str(len(rows_data)) + ' rows!','success')
    except Exception as e:
        try: conn.rollback()
        except: pass
        flash('Error: ' + str(e),'error')
    return redirect(url_for('main.index'))

# ==================== VIEW TABLE ====================
@blueprint.route('/table/<tablename>')
def view_table(tablename):
    if not session.get('user'): return redirect(url_for('main.login_page'))
    conn = get_connection(); cur = conn.cursor()
    cur.execute("SELECT table_name FROM _meta ORDER BY created_at DESC;"); tables = cur.fetchall()
    try:
        cur.execute('SELECT * FROM "' + tablename + '" LIMIT 1;')
        columns = [d[0] for d in cur.description if d[0] != 'id']
        cur.execute('SELECT COUNT(*) FROM "' + tablename + '";'); total_rows = cur.fetchone()[0]
        cur.execute('SELECT * FROM "' + tablename + '" ORDER BY id LIMIT 500;'); rows = cur.fetchall()
        numeric_columns = get_numeric_columns(cur, tablename)
        stats = _compute_stats(cur, tablename, columns, total_rows)
    except Exception as e: conn.rollback(); return "Error: " + str(e), 500
    cur.close(); conn.close()
    tabs = ''.join(['<a href="' + url_prefix + '/table/' + t[0] + '">' + t[0] + '</a>' for t in tables])
    rows_html = ''
    for i, row in enumerate(rows):
        rows_html += '<tr><td>' + str(i+1) + '</td>'
        for cell in row[1:]: rows_html += '<td>' + str(cell) + '</td>'
        rows_html += '<td><button class="btn btn-purple btn-sm" onclick="openPhotoModal(\'' + str(row[0]) + '\')">📷</button> <button class="btn btn-green btn-sm" onclick="openEditModal(\'' + str(row[0]) + '\',' + json.dumps(row) + ')">Edit</button> <a href="' + url_prefix + '/delete-row/' + tablename + '/' + str(row[0]) + '" class="btn btn-red btn-sm" onclick="return confirm(\'Delete?\')">Del</a></td></tr>'
    chart_cols = ''.join(['<option value="' + c + '">' + c + '</option>' for c in columns])
    chart_nums = ''.join(['<option value="' + c + '">' + c + '</option>' for c in numeric_columns])
    stats_html = ''.join(['<div class="stat-box"><div class="num" style="color:var(--blue);">' + str(s['value']) + '</div><div class="label">' + s['title'] + '</div></div>' for s in stats])
    
    html = BASE_STYLE + THEME_SCRIPT + hdr(session['user']) + '<div class="tabs">' + tabs + '</div>'
    html += '<div class="card"><div class="flex-between"><h3>' + tablename + ' <span class="badge">' + str(total_rows) + ' rows</span></h3>'
    html += '<div><button class="btn btn-green btn-sm" onclick="openAddModal()">+ Add</button> '
    html += '<a href="' + url_prefix + '/export/' + tablename + '?format=csv" class="btn btn-orange btn-sm">CSV</a> '
    html += '<a href="' + url_prefix + '/export/' + tablename + '?format=xlsx" class="btn btn-orange btn-sm">Excel</a> '
    html += '<a href="' + url_prefix + '/export/' + tablename + '?format=pdf" class="btn btn-purple btn-sm">PDF</a> '
    html += '<a href="' + url_prefix + '/copy-table/' + tablename + '" class="btn btn-primary btn-sm">Copy</a> '
    html += '<a href="' + url_prefix + '/do-rename/' + tablename + '" class="btn btn-primary btn-sm">Rename</a> '
    html += '<a href="' + url_prefix + '/truncate-table/' + tablename + '" class="btn btn-orange btn-sm" onclick="return confirm(\'Clear?\')">Clear</a> '
    html += '<a href="' + url_prefix + '/delete-table/' + tablename + '" class="btn btn-red btn-sm" onclick="return confirm(\'Delete?\')">Delete</a></div></div>'
    
    html += '<div class="filter-bar"><input type="text" id="searchInput" placeholder="Search..."><select id="filterColumn">' + chart_cols + '</select><input type="text" id="filterValue"><button class="btn btn-primary btn-sm" onclick="applyFilters()">Apply</button></div>'
    html += '<div class="scrollable"><table><tr><th>#</th>' + ''.join(['<th onclick="sortTable(\'' + c + '\')">' + c + '</th>' for c in columns]) + '<th>Actions</th></tr>' + rows_html + '</table></div></div>'
    
    html += '<div class="modal-overlay" id="editModal"><div class="modal"><h3 id="modalTitle">Edit</h3><form method="POST" id="modalForm"><input type="hidden" name="rowid"><div id="modalFields"></div><div class="btn-row"><button type="button" class="btn btn-red btn-sm" onclick="closeModal()">Cancel</button><button type="submit" class="btn btn-green btn-sm">Save</button></div></form></div></div>'
    html += '<div class="modal-overlay" id="photoModal"><div class="modal"><h3>Upload Photo</h3><form method="POST" action="' + url_prefix + '/upload-photo" enctype="multipart/form-data"><input type="hidden" name="table_name" value="' + tablename + '"><input type="hidden" name="row_id" id="photoRowId"><input type="file" name="photo" accept="image/*"><div class="btn-row"><button type="button" class="btn btn-red btn-sm" onclick="closePhotoModal()">Cancel</button><button type="submit" class="btn btn-green btn-sm">Upload</button></div></form></div></div>'
    
    html += '<div class="card"><h3>Quick Stats</h3><div class="row">' + stats_html + '</div></div>'
    html += '<div class="card"><h3>Chart</h3><select id="chartType" onchange="buildChart()"><option value="bar">Bar</option><option value="pie">Pie</option></select><select id="chartLabelCol">' + chart_cols + '</select><select id="chartValueCol">' + chart_nums + '</select><input type="color" id="chartColor" value="#58a6ff" onchange="buildChart()"><div class="chart-container"><canvas id="mainChart"></canvas></div></div>'
    html += '<div class="card"><h3>Reports</h3><div class="report-grid">'
    html += '<div><h4>Group By</h4><form method="POST" action="' + url_prefix + '/quick-report/' + tablename + '"><select name="groupcol">' + chart_cols + '</select><select name="aggfunc"><option value="COUNT">COUNT</option><option value="SUM">SUM</option></select><select name="aggcol"><option value="">-- Value --</option>' + chart_nums + '</select><button class="btn btn-green btn-sm" style="width:100%;margin-top:6px;">Run</button></form></div>'
    html += '<div><h4>Top N</h4><form method="POST" action="' + url_prefix + '/top-n/' + tablename + '"><select name="col">' + chart_cols + '</select><input type="number" name="n" value="10"><button class="btn btn-orange btn-sm" style="width:100%;margin-top:6px;">Show</button></form></div>'
    html += '<div><h4>Duplicates</h4><form method="POST" action="' + url_prefix + '/duplicates/' + tablename + '"><select name="col">' + chart_cols + '</select><button class="btn btn-red btn-sm" style="width:100%;margin-top:6px;">Find</button></form></div>'
    html += '<div><h4>Null Analysis</h4><form method="POST" action="' + url_prefix + '/nulls/' + tablename + '"><button class="btn btn-purple btn-sm" style="width:100%;">Analyze</button></form></div>'
    html += '</div></div>'
    html += '<div class="card"><h3>SQL Console</h3><form method="POST" action="' + url_prefix + '/sql"><textarea name="query" rows="3"></textarea><button class="btn btn-primary" style="margin-top:8px;">Execute</button></form></div>'
    
    html += '<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"> </script>'
    html += '<script>const columns=' + json.dumps(columns) + ';const allRows=' + json.dumps([dict(zip(columns, row[1:])) for row in rows]) + ';let mainChart=null;'
    html += 'function buildChart(){if(!allRows.length)return;const t=document.getElementById("chartType").value,l=document.getElementById("chartLabelCol").value,v=document.getElementById("chartValueCol").value,c=document.getElementById("chartColor").value;if(!l||!v)return;const g={};allRows.forEach(r=>{const k=String(r[l]||"(empty)"),val=parseFloat(String(r[v]).replace(",","."))||0;g[k]=(g[k]||0)+val});const e=Object.entries(g).sort((a,b)=>b[1]-a[1]).slice(0,20),lbs=e.map(x=>x[0]),d=e.map(x=>x[1]),ctx=document.getElementById("mainChart").getContext("2d");if(mainChart)mainChart.destroy();mainChart=new Chart(ctx,{type:t,data:{labels:lbs,datasets:[{label:v,data:d,backgroundColor:t==="pie"?lbs.map((_,i)=>`hsl(${i*360/lbs.length},70%,50%)`):c+"99",borderColor:c}]},options:{responsive:!0,maintainAspectRatio:!1}})}buildChart();'
    html += 'function applyFilters(){const s=document.getElementById("searchInput").value.toLowerCase(),col=document.getElementById("filterColumn").value,val=document.getElementById("filterValue").value.toLowerCase();document.querySelectorAll("table tr:not(:first-child)").forEach(tr=>{let show=!0;if(s)show=tr.textContent.toLowerCase().includes(s);if(show&&col&&val)show=tr.children[columns.indexOf(col)+1]?.textContent.toLowerCase().includes(val);tr.style.display=show?"":"none"})}'
    html += 'let sortDir={};function sortTable(col){sortDir[col]=!sortDir[col];const idx=columns.indexOf(col)+1,tbody=document.querySelector("table"),rows=Array.from(tbody.querySelectorAll("tr:not(:first-child)"));rows.sort((a,b)=>{const va=a.children[idx]?.textContent.trim()||"",vb=b.children[idx]?.textContent.trim()||"",na=parseFloat(va),nb=parseFloat(vb);if(!isNaN(na)&&!isNaN(nb))return sortDir[col]?na-nb:nb-na;return sortDir[col]?va.localeCompare(vb):vb.localeCompare(va)});rows.forEach(r=>tbody.appendChild(r))}'
    html += 'function openAddModal(){document.getElementById("modalTitle").textContent="Add Row";document.getElementById("modalForm").action="' + url_prefix + '/add-row/' + tablename + '";let h="";columns.forEach(c=>{h+="<label>"+c+"</label><input type=text name="+c+">"});document.getElementById("modalFields").innerHTML=h;document.getElementById("editModal").classList.add("active")}'
    html += 'function openEditModal(id,rd){document.getElementById("modalTitle").textContent="Edit #"+id;document.getElementById("modalForm").action="' + url_prefix + '/edit-row/' + tablename + '/"+id;let h="";columns.forEach((c,i)=>{h+="<label>"+c+"</label><input type=text name="+c+" value=\\""+String(rd[i+1]||"").replace(/"/g,"&quot;")+"\\">"});document.getElementById("modalFields").innerHTML=h;document.getElementById("editModal").classList.add("active")}'
    html += 'function closeModal(){document.getElementById("editModal").classList.remove("active")}'
    html += 'function openPhotoModal(rowId){document.getElementById("photoRowId").value=rowId;document.getElementById("photoModal").classList.add("active")}'
    html += 'function closePhotoModal(){document.getElementById("photoModal").classList.remove("active")}'
    html += '</script></div></body></html>'
    return html

# ==================== PHOTO UPLOAD ====================
@blueprint.route('/upload-photo', methods=['POST'])
def upload_photo():
    if not session.get('user'): return redirect(url_for('main.login_page'))
    table_name = request.form.get('table_name',''); row_id = request.form.get('row_id','')
    file = request.files.get('photo')
    if file and file.filename:
        ext = os.path.splitext(file.filename)[1]
        filename = str(uuid.uuid4())[:8] + ext
        file.save(os.path.join(UPLOAD_FOLDER, filename))
        conn = get_connection(); cur = conn.cursor()
        cur.execute("INSERT INTO _photos (table_name, row_id, filename) VALUES (%s,%s,%s)", (table_name, int(row_id), filename))
        conn.commit(); cur.close(); conn.close()
        flash('Photo uploaded!','success')
    return redirect(url_for('main.view_table', tablename=table_name))

@blueprint.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)

# ==================== CRUD ====================
@blueprint.route('/add-row/<tablename>', methods=['POST'])
def add_row(tablename):
    if not session.get('user'): return redirect(url_for('main.login_page'))
    conn = get_connection(); cur = conn.cursor()
    cur.execute('SELECT * FROM "' + tablename + '" LIMIT 1;'); columns = [d[0] for d in cur.description if d[0] != 'id']
    vals = [request.form.get(c,'') for c in columns]
    cur.execute('INSERT INTO "' + tablename + '" (' + ','.join(['"' + c + '"' for c in columns]) + ') VALUES (' + ','.join(['%s']*len(columns)) + ');', vals)
    conn.commit(); cur.close(); conn.close(); log_action(session['user'], 'ADD_ROW', tablename)
    flash('Row added!','success'); return redirect(url_for('main.view_table', tablename=tablename))

@blueprint.route('/edit-row/<tablename>/<int:rowid>', methods=['POST'])
def edit_row(tablename, rowid):
    if not session.get('user'): return redirect(url_for('main.login_page'))
    conn = get_connection(); cur = conn.cursor()
    cur.execute('SELECT * FROM "' + tablename + '" LIMIT 1;'); columns = [d[0] for d in cur.description if d[0] != 'id']
    sets = ', '.join(['"' + c + '"=%s' for c in columns]); vals = [request.form.get(c,'') for c in columns] + [rowid]
    cur.execute('UPDATE "' + tablename + '" SET ' + sets + ' WHERE id=%s;', vals)
    conn.commit(); cur.close(); conn.close(); log_action(session['user'], 'EDIT_ROW', tablename, 'id='+str(rowid))
    flash('Row updated!','success'); return redirect(url_for('main.view_table', tablename=tablename))

@blueprint.route('/delete-row/<tablename>/<int:rowid>')
def delete_row(tablename, rowid):
    if not session.get('user'): return redirect(url_for('main.login_page'))
    conn = get_connection(); cur = conn.cursor()
    cur.execute('DELETE FROM "' + tablename + '" WHERE id=%s;', (rowid,)); conn.commit()
    cur.close(); conn.close(); log_action(session['user'], 'DELETE_ROW', tablename, 'id='+str(rowid))
    flash('Row deleted!','success'); return redirect(url_for('main.view_table', tablename=tablename))

# ==================== TABLE OPS ====================
@blueprint.route('/copy-table/<tablename>')
def copy_table(tablename):
    new_name = sanitize_name(tablename + '_copy')
    conn = get_connection(); cur = conn.cursor()
    cur.execute('CREATE TABLE "' + new_name + '" AS SELECT * FROM "' + tablename + '";')
    cur.execute("INSERT INTO _meta (table_name) VALUES (%s) ON CONFLICT DO NOTHING;", (new_name,)); conn.commit()
    cur.close(); conn.close(); log_action(session['user'], 'COPY_TABLE', tablename, '->'+new_name)
    flash('Copied to "' + new_name + '"!','success'); return redirect(url_for('main.view_table', tablename=new_name))

@blueprint.route('/do-rename/<tablename>', methods=['GET','POST'])
def do_rename(tablename):
    if request.method == 'GET': return BASE_STYLE + hdr(session['user']) + '<div class="card"><h3>Rename: ' + tablename + '</h3><form method="POST"><input type="text" name="new_name" required><button class="btn btn-primary" style="margin-top:12px;">Rename</button></form></div></div></body></html>'
    new_name = sanitize_name(request.form.get('new_name',''))
    conn = get_connection(); cur = conn.cursor()
    cur.execute('ALTER TABLE "' + tablename + '" RENAME TO "' + new_name + '";')
    cur.execute("UPDATE _meta SET table_name=%s WHERE table_name=%s;", (new_name, tablename)); conn.commit()
    cur.close(); conn.close(); log_action(session['user'], 'RENAME_TABLE', tablename, '->'+new_name)
    flash('Renamed!','success'); return redirect(url_for('main.view_table', tablename=new_name))

@blueprint.route('/truncate-table/<tablename>')
def truncate_table(tablename):
    conn = get_connection(); cur = conn.cursor()
    cur.execute('TRUNCATE TABLE "' + tablename + '";'); conn.commit()
    cur.close(); conn.close(); log_action(session['user'], 'TRUNCATE', tablename)
    flash('Cleared!','success'); return redirect(url_for('main.view_table', tablename=tablename))

@blueprint.route('/delete-table/<tablename>')
def delete_table(tablename):
    conn = get_connection(); cur = conn.cursor()
    cur.execute('DROP TABLE IF EXISTS "' + tablename + '";'); cur.execute("DELETE FROM _meta WHERE table_name=%s;",(tablename,)); conn.commit()
    cur.close(); conn.close(); log_action(session['user'], 'DELETE_TABLE', tablename)
    flash('Deleted!','success'); return redirect(url_for('main.index'))

# ==================== OVERVIEW ====================
@blueprint.route('/overview')
def overview():
    if not session.get('user'): return redirect(url_for('main.login_page'))
    conn = get_connection(); cur = conn.cursor()
    cur.execute("SELECT table_name FROM _meta ORDER BY created_at DESC;"); tables = cur.fetchall()
    info = []
    for t in tables:
        try:
            cur.execute('SELECT COUNT(*) FROM "' + t[0] + '";'); rc = cur.fetchone()[0]
            cur.execute('SELECT * FROM "' + t[0] + '" LIMIT 1;'); cc = len([d for d in cur.description if d[0]!='id'])
            info.append((t[0], rc, cc))
        except: info.append((t[0], 0, 0))
    cur.close(); conn.close()
    rows_html = ''.join(['<tr><td><a href="' + url_prefix + '/table/' + i[0] + '">' + i[0] + '</a></td><td>' + str(i[1]) + '</td><td>' + str(i[2]) + '</td></tr>' for i in info])
    total_rows = sum(i[1] for i in info)
    return BASE_STYLE + THEME_SCRIPT + hdr(session['user']) + '<div class="card"><h3>Overview</h3><div class="row"><div class="stat-box"><div class="num">' + str(len(tables)) + '</div><div class="label">Tables</div></div><div class="stat-box"><div class="num">' + str(total_rows) + '</div><div class="label">Total Rows</div></div></div></div><div class="card"><table><tr><th>Table</th><th>Rows</th><th>Columns</th></tr>' + rows_html + '</table></div></div></body></html>'

# ==================== DASHBOARD ====================
@blueprint.route('/dashboard')
def dashboard():
    if not session.get('user'): return redirect(url_for('main.login_page'))
    t1 = request.args.get('t1',''); t2 = request.args.get('t2','')
    conn = get_connection(); cur = conn.cursor()
    cur.execute("SELECT table_name FROM _meta ORDER BY created_at DESC;"); tables = cur.fetchall()
    charts_js = ""; stats_all = []
    if t1:
        try:
            cur.execute('SELECT * FROM "' + t1 + '" LIMIT 1;'); cols = [d[0] for d in cur.description if d[0]!='id']
            cur.execute('SELECT COUNT(*) FROM "' + t1 + '";'); tr = cur.fetchone()[0]
            stats_all += _compute_stats(cur, t1, cols, tr)
            cur.execute('SELECT "' + cols[0] + '", COUNT(*) FROM "' + t1 + '" GROUP BY "' + cols[0] + '" ORDER BY 2 DESC LIMIT 15;')
            r = cur.fetchall(); labels = json.dumps([str(x[0]) for x in r]); values = json.dumps([x[1] for x in r])
            charts_js += f'<div class="chart-container"><canvas id="chart1"></canvas></div><script>new Chart(document.getElementById("chart1"),{{type:"bar",data:{{labels:{labels},datasets:[{{label:"{t1}",data:{values},backgroundColor:"#58a6ff99"}}]}},options:{{responsive:!0,maintainAspectRatio:!1}}}});</script>'
        except: pass
    if t2:
        try:
            cur.execute('SELECT * FROM "' + t2 + '" LIMIT 1;'); cols = [d[0] for d in cur.description if d[0]!='id']
            cur.execute('SELECT COUNT(*) FROM "' + t2 + '";'); tr = cur.fetchone()[0]
            stats_all += _compute_stats(cur, t2, cols, tr)
            cur.execute('SELECT "' + cols[0] + '", COUNT(*) FROM "' + t2 + '" GROUP BY "' + cols[0] + '" ORDER BY 2 DESC LIMIT 15;')
            r = cur.fetchall(); labels = json.dumps([str(x[0]) for x in r]); values = json.dumps([x[1] for x in r])
            charts_js += f'<div class="chart-container"><canvas id="chart2"></canvas></div><script>new Chart(document.getElementById("chart2"),{{type:"bar",data:{{labels:{labels},datasets:[{{label:"{t2}",data:{values},backgroundColor:"#3fb95099"}}]}},options:{{responsive:!0,maintainAspectRatio:!1}}}});</script>'
        except: pass
    cur.close(); conn.close()
    table_opts = ''.join(['<option value="' + t[0] + '">' + t[0] + '</option>' for t in tables])
    stats_html = ''.join(['<div class="stat-box"><div class="num" style="color:var(--blue);">' + str(s['value']) + '</div><div class="label">' + s['title'] + '</div></div>' for s in stats_all])
    return BASE_STYLE + THEME_SCRIPT + hdr(session['user']) + '<div class="card"><h3>Dashboard 2.0</h3><form class="row"><select name="t1"><option value="">-- Table 1 --</option>' + table_opts + '</select><select name="t2"><option value="">-- Table 2 --</option>' + table_opts + '</select><button class="btn btn-primary">Load</button></form></div><div class="dashboard-grid">' + charts_js + '</div><div class="card"><h3>Stats</h3><div class="row">' + stats_html + '</div></div><script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"> </script></div></body></html>'

# ==================== PIVOT ====================
@blueprint.route('/pivot', methods=['GET','POST'])
def pivot():
    if not session.get('user'): return redirect(url_for('main.login_page'))
    conn = get_connection(); cur = conn.cursor()
    cur.execute("SELECT table_name FROM _meta ORDER BY created_at DESC;"); tables = cur.fetchall()
    pivot_html = ""; table_opts = ''.join(['<option value="' + t[0] + '">' + t[0] + '</option>' for t in tables])
    if request.method == 'POST':
        tn = request.form.get('table_name',''); rc = request.form.get('row_col',''); cc = request.form.get('col_col',''); vc = request.form.get('val_col',''); agg = request.form.get('agg','COUNT')
        if tn and rc and cc and vc:
            try:
                cur.execute('SELECT "' + rc + '", "' + cc + '", ' + agg + '("' + vc + '"::numeric) FROM "' + tn + '" GROUP BY "' + rc + '", "' + cc + '" ORDER BY "' + rc + '";')
                raw = cur.fetchall(); rv = sorted(set(r[0] for r in raw)); cv = sorted(set(r[1] for r in raw))
                matrix = {r:{c:0 for c in cv} for r in rv}
                for r in raw: matrix[r[0]][r[1]] = round(r[2],2) if isinstance(r[2],float) else r[2]
                pivot_html = '<div class="scrollable"><table><tr><th>' + rc + '</th>' + ''.join('<th>' + c + '</th>' for c in cv) + '</tr>'
                for r_val in rv: pivot_html += '<tr><td><b>' + str(r_val) + '</b></td>' + ''.join('<td>' + str(matrix[r_val][c]) + '</td>' for c in cv) + '</tr>'
                pivot_html += '</table></div>'
            except Exception as e: pivot_html = '<p style="color:var(--red);">Error: ' + str(e) + '</p>'
    cur.close(); conn.close()
    return BASE_STYLE + THEME_SCRIPT + hdr(session['user']) + '<div class="card"><h3>Pivot Table</h3><form method="POST" class="row"><select name="table_name"><option value="">-- Table --</option>' + table_opts + '</select><input name="row_col" placeholder="Row"><input name="col_col" placeholder="Column"><input name="val_col" placeholder="Value"><select name="agg"><option value="COUNT">COUNT</option><option value="SUM">SUM</option><option value="AVG">AVG</option></select><button class="btn btn-primary">Build</button></form></div>' + pivot_html + '</div></body></html>'

# ==================== COMPARE ====================
@blueprint.route('/compare', methods=['GET','POST'])
def compare():
    if not session.get('user'): return redirect(url_for('main.login_page'))
    conn = get_connection(); cur = conn.cursor()
    cur.execute("SELECT table_name FROM _meta ORDER BY created_at DESC;"); tables = cur.fetchall()
    diff_html = ""; table_opts = ''.join(['<option value="' + t[0] + '">' + t[0] + '</option>' for t in tables])
    if request.method == 'POST':
        t1 = request.form.get('t1',''); t2 = request.form.get('t2','')
        if t1 and t2:
            try:
                cur.execute('SELECT COUNT(*) FROM "' + t1 + '";'); c1 = cur.fetchone()[0]
                cur.execute('SELECT COUNT(*) FROM "' + t2 + '";'); c2 = cur.fetchone()[0]
                cur.execute('SELECT * FROM "' + t1 + '" LIMIT 1;'); cols1 = set(d[0] for d in cur.description if d[0]!='id')
                cur.execute('SELECT * FROM "' + t2 + '" LIMIT 1;'); cols2 = set(d[0] for d in cur.description if d[0]!='id')
                diff_html = f'<p><b>Rows:</b> {t1}: {c1} | {t2}: {c2} | Diff: {abs(c1-c2)}</p>'
                if cols1 != cols2: diff_html += f'<p><b>Columns only in {t1}:</b> {cols1-cols2}</p><p><b>Columns only in {t2}:</b> {cols2-cols1}</p>'
                else: diff_html += '<p><b>Columns:</b> Identical</p>'
            except Exception as e: diff_html = '<p style="color:var(--red);">Error: ' + str(e) + '</p>'
    cur.close(); conn.close()
    return BASE_STYLE + THEME_SCRIPT + hdr(session['user']) + '<div class="card"><h3>Compare Tables</h3><form method="POST" class="row"><select name="t1"><option value="">-- Table 1 --</option>' + table_opts + '</select><select name="t2"><option value="">-- Table 2 --</option>' + table_opts + '</select><button class="btn btn-primary">Compare</button></form></div>' + diff_html + '</div></body></html>'

# ==================== TEMPLATES ====================
@blueprint.route('/templates')
def templates():
    if not session.get('user'): return redirect(url_for('main.login_page'))
    conn = get_connection(); cur = conn.cursor()
    cur.execute("SELECT * FROM _templates ORDER BY created_at DESC;"); temps = cur.fetchall()
    cur.close(); conn.close()
    rows = ''.join(['<tr><td>' + t[1] + '</td><td>' + t[2] + '</td><td><a href="' + url_prefix + '/run-template/' + str(t[0]) + '" class="btn btn-green btn-sm">Run</a> <a href="' + url_prefix + '/delete-template/' + str(t[0]) + '" class="btn btn-red btn-sm" onclick="return confirm(\'Delete?\')">Del</a></td></tr>' for t in temps])
    return BASE_STYLE + THEME_SCRIPT + hdr(session['user']) + '<div class="card"><h3>Templates</h3><table><tr><th>Name</th><th>Table</th><th>Actions</th></tr>' + rows + '</table></div></body></html>'

@blueprint.route('/save-template', methods=['POST'])
def save_template():
    if not session.get('user'): return redirect(url_for('login_page'))
    conn = get_connection(); cur = conn.cursor()
    cur.execute("INSERT INTO _templates (name, table_name, config) VALUES (%s,%s,%s)", (request.form.get('template_name'), request.form.get('table_name'), request.form.get('config'))); conn.commit()
    cur.close(); conn.close(); flash('Saved!','success'); return redirect(url_for('main.templates'))

@blueprint.route('/run-template/<int:tid>')
def run_template(tid):
    conn = get_connection(); cur = conn.cursor(); cur.execute("SELECT table_name FROM _templates WHERE id=%s",(tid,)); t = cur.fetchone(); cur.close(); conn.close()
    return redirect(url_for('main.view_table', tablename=t[0])) if t else ('Not found', 404)

@blueprint.route('/delete-template/<int:tid>')
def delete_template(tid):
    conn = get_connection(); cur = conn.cursor(); cur.execute("DELETE FROM _templates WHERE id=%s",(tid,)); conn.commit()
    cur.close(); conn.close(); flash('Deleted!','success'); return redirect(url_for('main.templates'))

# ==================== LOGS ====================
@blueprint.route('/logs')
def logs():
    if not session.get('user'): return redirect(url_for('main.login_page'))
    conn = get_connection(); cur = conn.cursor()
    cur.execute("SELECT * FROM _logs ORDER BY created_at DESC LIMIT 200;"); logs = cur.fetchall()
    cur.close(); conn.close()
    rows = ''.join(['<tr><td>' + str(l[5]) + '</td><td>' + l[1] + '</td><td>' + l[2] + '</td><td>' + l[3] + '</td><td>' + (l[4] or '')[:100] + '</td></tr>' for l in logs])
    return BASE_STYLE + THEME_SCRIPT + hdr(session['user']) + '<div class="card"><h3>Activity Logs</h3><div class="scrollable"><table><tr><th>Time</th><th>User</th><th>Action</th><th>Table</th><th>Details</th></tr>' + rows + '</table></div></div></body></html>'

# ==================== SQL ====================
@blueprint.route('/sql', methods=['POST'])
def sql_query():
    if not session.get('user'): return redirect(url_for('main.login_page'))
    query = request.form.get('query',''); conn = get_connection(); cur = conn.cursor(); cols = rows = err = None
    try:
        cur.execute(query)
        if cur.description: cols = [d[0] for d in cur.description]; rows = cur.fetchall()
        else: conn.commit(); flash('Executed!','success')
    except Exception as e: err = str(e)
    cur.close(); conn.close()
    html = ''
    if cols: html = '<table><tr>' + ''.join('<th>' + c + '</th>' for c in cols) + '</tr>' + ''.join('<tr>' + ''.join('<td>' + str(c) + '</td>' for c in r) + '</tr>' for r in rows) + '</table>'
    if err: html += '<div class="flash flash-error">' + err + '</div>'
    return BASE_STYLE + hdr(session['user']) + '<div class="card"><h3>SQL Result</h3>' + html + '</div></body></html>'

# ==================== EXPORT ====================
@blueprint.route('/export/<tablename>')
def export_table(tablename):
    fmt = request.args.get('format','csv'); conn = get_connection(); cur = conn.cursor()
    cur.execute('SELECT * FROM "' + tablename + '";'); cols = [d[0] for d in cur.description if d[0]!='id']; rows = cur.fetchall()
    cur.close(); conn.close()
    if fmt == 'xlsx':
        wb = Workbook(); ws = wb.active; ws.append(cols)
        for r in rows: ws.append(r[1:])
        out = io.BytesIO(); wb.save(out); out.seek(0)
        return Response(out.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    elif fmt == 'pdf':
        buf = io.BytesIO(); doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
        data = [cols] + [[str(c) for c in r[1:]] for r in rows]
        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#2196F3')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('FONTSIZE',(0,0),(-1,-1),7),('GRID',(0,0),(-1,-1),0.5,colors.grey)]))
        doc.build([t]); buf.seek(0)
        return Response(buf.getvalue(), mimetype='application/pdf')
    else:
        out = io.StringIO(); w = csv.writer(out); w.writerow(cols)
        for r in rows: w.writerow(r[1:])
        return Response(out.getvalue(), mimetype='text/csv')

# ==================== API ====================
@blueprint.route('/api/v1/tables')
@api_required
def api_tables():
    conn = get_connection(); cur = conn.cursor()
    cur.execute("SELECT table_name FROM _meta;"); tables = [t[0] for t in cur.fetchall()]
    cur.close(); conn.close()
    return Response(json.dumps({"tables": tables}), mimetype='application/json')

@blueprint.route('/api/v1/table/<tablename>')
@api_required
def api_table_data(tablename):
    conn = get_connection(); cur = conn.cursor()
    try:
        cur.execute('SELECT * FROM "' + tablename + '" LIMIT 1000;')
        cols = [d[0] for d in cur.description if d[0]!='id']; rows = cur.fetchall()
        data = [dict(zip(cols, r[1:])) for r in rows]
    except: data = []
    cur.close(); conn.close()
    return Response(json.dumps({"table": tablename, "rows": data}, default=str), mimetype='application/json')

# ==================== BACKUP ====================
@blueprint.route('/backup')
def backup():
    if not session.get('user'): return redirect(url_for('main.login_page'))
    conn = get_connection(); cur = conn.cursor()
    cur.execute("SELECT table_name FROM _meta;"); tables = [t[0] for t in cur.fetchall()]
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for t in tables:
            cur.execute('SELECT * FROM "' + t + '";'); cols = [d[0] for d in cur.description if d[0]!='id']; rows = cur.fetchall()
            out = io.StringIO(); w = csv.writer(out); w.writerow(cols)
            for r in rows: w.writerow(r[1:]); zf.writestr(t + '.csv', out.getvalue())
    cur.close(); conn.close(); buf.seek(0)
    return Response(buf.getvalue(), mimetype='application/zip', headers={"Content-Disposition": "attachment;filename=backup_" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".zip"})

# ==================== REPORTS ====================
@blueprint.route('/quick-report/<tablename>', methods=['POST'])
def quick_report(tablename):
    if not session.get('user'): return redirect(url_for('main.login_page'))
    groupcol = request.form.get('groupcol',''); aggcol = request.form.get('aggcol',''); aggfunc = request.form.get('aggfunc','COUNT')
    conn = get_connection(); cur = conn.cursor()
    q = 'SELECT "' + groupcol + '", ' + aggfunc + '("' + aggcol + '"::numeric) FROM "' + tablename + '" GROUP BY "' + groupcol + '" ORDER BY 2 DESC LIMIT 50;' if aggcol else 'SELECT "' + groupcol + '", COUNT(*) FROM "' + tablename + '" GROUP BY "' + groupcol + '" ORDER BY 2 DESC LIMIT 50;'
    cur.execute(q); rows = cur.fetchall(); cols = [d[0] for d in cur.description]; cur.close(); conn.close()
    return BASE_STYLE + hdr(session['user']) + '<div class="card"><h3>Report</h3><table><tr>' + ''.join('<th>' + c + '</th>' for c in cols) + '</tr>' + ''.join('<tr>' + ''.join('<td>' + str(c) + '</td>' for c in r) + '</tr>' for r in rows) + '</table></div></body></html>'

@blueprint.route('/top-n/<tablename>', methods=['POST'])
def top_n(tablename):
    if not session.get('user'): return redirect(url_for('main.login_page'))
    col = request.form.get('col',''); n = int(request.form.get('n',10))
    conn = get_connection(); cur = conn.cursor()
    cur.execute('SELECT "' + col + '", COUNT(*) AS cnt FROM "' + tablename + '" GROUP BY "' + col + '" ORDER BY cnt DESC LIMIT ' + str(n) + ';')
    rows = cur.fetchall(); cols = [d[0] for d in cur.description]; cur.close(); conn.close()
    return BASE_STYLE + hdr(session['user']) + '<div class="card"><h3>Top ' + str(n) + '</h3><table><tr>' + ''.join('<th>' + c + '</th>' for c in cols) + '</tr>' + ''.join('<tr>' + ''.join('<td>' + str(c) + '</td>' for c in r) + '</tr>' for r in rows) + '</table></div></body></html>'

@blueprint.route('/duplicates/<tablename>', methods=['POST'])
def duplicates(tablename):
    if not session.get('user'): return redirect(url_for('main.login_page'))
    col = request.form.get('col','')
    conn = get_connection(); cur = conn.cursor()
    cur.execute('SELECT "' + col + '", COUNT(*) AS cnt FROM "' + tablename + '" GROUP BY "' + col + '" HAVING COUNT(*) > 1 ORDER BY cnt DESC LIMIT 100;')
    rows = cur.fetchall(); cols = [d[0] for d in cur.description]; cur.close(); conn.close()
    return BASE_STYLE + hdr(session['user']) + '<div class="card"><h3>Duplicates</h3><table><tr>' + ''.join('<th>' + c + '</th>' for c in cols) + '</tr>' + ''.join('<tr>' + ''.join('<td>' + str(c) + '</td>' for c in r) + '</tr>' for r in rows) + '</table></div></body></html>'

@blueprint.route('/nulls/<tablename>', methods=['POST'])
def nulls(tablename):
    if not session.get('user'): return redirect(url_for('main.login_page'))
    conn = get_connection(); cur = conn.cursor()
    cur.execute('SELECT * FROM "' + tablename + '" LIMIT 1;'); columns = [d[0] for d in cur.description if d[0]!='id']
    cur.execute('SELECT COUNT(*) FROM "' + tablename + '";'); total_rows = cur.fetchone()[0]
    result = []; headers = ['Column', 'Empty', 'Filled', 'Fill %']
    for col in columns:
        cur.execute('SELECT COUNT(*) FROM "' + tablename + '" WHERE "' + col + '" IS NULL OR "' + col + '" = \'\';')
        nulls_c = cur.fetchone()[0]; result.append([col, nulls_c, total_rows - nulls_c, str(round(100*(total_rows-nulls_c)/total_rows,1))+'%'])
    cur.close(); conn.close()
    return BASE_STYLE + hdr(session['user']) + '<div class="card"><h3>Null Analysis</h3><table><tr>' + ''.join('<th>' + h + '</th>' for h in headers) + '</tr>' + ''.join('<tr>' + ''.join('<td>' + str(c) + '</td>' for c in r) + '</tr>' for r in result) + '</table></div></body></html>'

app.register_blueprint(blueprint)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
